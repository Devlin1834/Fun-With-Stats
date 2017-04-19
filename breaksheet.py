# -*- coding: utf-8 -*-
"""
Created on Sun Apr 16 21:58:28 2017

@author: Devlin
"""

import PyPDF2 as pdf
import re
import openpyxl as xl
import random as rn
from openpyxl.styles import Border

###################################################################################################
####################################### Regular Expressions #######################################

shift_regex = re.compile(r'''((\d)?\d:\d\d\D) - ((\d)?\d:\d\d\D)''')

###################################################################################################
####################################### Functions #################################################


###################################################################################################
####################################### Lets Make a Sheet #########################################

brk = xl.Workbook()
sheet = brk.active
sheet.title = 'Test'

wide = ['A', 'E', 'I']
thin = ['B', 'C', 'F', 'G', 'K']
bord = ['D', 'H']
merge = ['A2:C2', 
         'E2:G2', 
         'I2:K2', 
         'A11:C11', 
         'E11:G11', 
         'A19:C19', 
         'E19:G19', 
         'A30:C30', 
         'E26:G26',
         'E33:G33',
         'I35:K35',
         'A42:C42',
         'E42:G42',
         'I40:K40'
         ]

happy_thoughts = ['Have a great day everybody!',
                  'Your smile makes my day!',
                  'May your guests be happy and your spoilage be low!',
                  'You\'re my favorite Team Captain!',
                  'There is always a silver lining!',
                  'Life ... uh ...finds a way',
                  'Respect is earned but courtesy is given',
                  'Shit happens, no need for blame, just learning',
                  'The journey is more important the destination',
                  'You create your own purpose in life',
                  'It can always get worse',
                  'Everyone wants to be happy'
                  ]

names = {'A2': 'Kong Mess Tent',
         'I2': 'Pull List',
         'E2': 'JP Turkey',
         'A11': 'JP Fruit',
         'E11': 'JP Beer',
         'A19': 'Pizza Predatoria',
         'E19': 'Pizza Predatoria Mini Grill',
         'E33': 'Stockers',
         'A30': 'Pred Cooks',
         'E42': 'Bussers',
         'I40': 'Team Captains',
         'E26': 'Grill Cooks',
         'A42': 'West Carts Cooks',
         'I35': 'Breakers'
         }

hours = ['A3:A4', 'E3:E4', 'A12:A13', 'E12:E13', 'A20:A21', 'E20:E21']

bold_name = xl.styles.Font(bold = True)

white_fill = xl.styles.PatternFill(start_color = 'FFFFFF',
                                   end_color = 'FFFFFF',
                                   fill_type = 'solid')

for w in wide:
    sheet.column_dimensions[w].width = 14.57
for t in thin:
    sheet.column_dimensions[t].width = 5.71
for b in bord:
    sheet.column_dimensions[b].width = 2.43    
for m in merge:
    sheet.merge_cells(m)
for h in hours:
    sheet.merge_cells(h)
for k in names.keys():
    sheet[k] = names[k]
for a in ['A3', 'E3', 'A12', 'E12', 'A20', 'E20']:
    sheet[a] = 'Cashier'
for o in ['B3', 'F3', 'B12', 'F12', 'B20', 'F20']:
    sheet[o] = 'Open'
for c in ['C3', 'G3', 'C12', 'G12', 'C20', 'G20']:
    sheet[c] = 'Close'
for b in names.keys():
    sheet[b].font = bold_name
    
sheet['I3'] = 'Name'
sheet['J3'] = 'Location'
sheet['K3'] = 'Time'
sheet.merge_cells('A1:K1')
sheet.merge_cells('A48:G49')

sheet['A48'] = happy_thoughts[rn.randint(1,len(happy_thoughts)-1)]
sheet['A48'].alignment = xl.styles.Alignment(wrapText = True)


for col in sheet.columns:
    for cell in col:
        alignment_obj = cell.alignment.copy(horizontal = 'center', vertical = 'center')
        cell.alignment = alignment_obj
        cell.fill = white_fill

sheet.page_margins.left = 0.25
sheet.page_margins.right = 0.25
sheet.page_margins.top = 0.25
sheet.page_margins.bottom = 0.25

brk.save('break.xlsx')

###################################################################################################
###################################### Range Deffinition ##########################################

cash_master = []
pull_master = []

kong_names = ['A5', 'A6', 'A7', 'A8', 'A9']
kong_in = ['B5', 'B6', 'B7', 'B8', 'B9']
kong_out = ['C5', 'C6', 'C7', 'C8', 'C9']

turkey_names = ['E5', 'E6', 'E7', 'E8', 'E9']
turkey_in = ['F5', 'F6', 'F7', 'F8', 'F9']
turkey_out = ['G5', 'G6', 'G7', 'G8', 'G9']

fruit_names = ['A15', 'A16', 'A17', 'A18', 'A19']
fruit_in = ['B15', 'B16', 'B17', 'B18', 'B19']
fruit_out = ['C15', 'C16', 'C17', 'C18', 'C19']

beer_names = ['E15', 'E16', 'E17', 'E18', 'E19']
beer_in = ['F15', 'F16', 'F17', 'F18', 'F19']
beer_out = ['G15', 'G16', 'G17', 'G18', 'G19']

pred_names = []
pred_in = []
pred_out = []

grill_names = []
grill_in = []
grill_out = []

stockers_names = []
stockers_in = []
stockers_out = []

pcooks_names = []
pcooks_in = []
pcooks_out = []

wcooks_names = []
wcooks_in = []
wcooks_out = []

gcooks_names = []
gcooks_in = []
gcooks_out = []

bussers_names = []
bussers_in = []
bussers_out = []

tc_names = []
tc_in = []
tc_out = []

###################################################################################################

OpenReport = open('OpenReport.pdf', 'rb')

sheet = pdf.PdfFileReader(OpenReport)

page1 = sheet.getPage(0)
page2 = sheet.getPage(1)

text1 = page1.extractText()

res = shift_regex.findall(text1)

print(res)
print(res[0][0])

